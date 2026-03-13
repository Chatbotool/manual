import streamlit as st
import google.generativeai as genai
import cv2
import os
import json
import time
import tempfile
import re
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill

# --- 初期設定 ---
st.set_page_config(page_title="AIマニュアル生成", layout="centered")

# Renderの環境変数からAPIキーを読み込む
API_KEY = os.environ.get("GEMINI_API_KEY")

# --- 画像切り出し関数 ---
def extract_frame(video_path, time_str, output_path):
    """指定された時間のフレームを動画から切り出して保存する"""
    try:
        minutes, seconds = map(float, time_str.split(':'))
        total_seconds = minutes * 60 + seconds
        cap = cv2.VideoCapture(video_path)
        fps = cap.get(cv2.CAP_PROP_FPS)
        cap.set(cv2.CAP_PROP_POS_FRAMES, int(total_seconds * fps))
        ret, frame = cap.read()
        if ret:
            cv2.imwrite(output_path, frame)
        cap.release()
        return ret
    except Exception:
        return False

# --- AI出力からJSON（辞書型）を安全に抽出する関数 ---
def extract_json_from_text(text):
    """AIが回答に余計な装飾を付けてもJSON部分( { } で囲まれた部分 )だけを抜き出す"""
    match = re.search(r'\{.*\}', text, re.DOTALL)
    if match:
        json_str = match.group(0)
        return json.loads(json_str)
    else:
        raise ValueError("AIの回答からJSON形式のデータを抽出できませんでした。")

# --- 画面UI構成 ---
st.title("🎥 AIマニュアル自動作成ツール")
st.write("動画をアップロードするだけで、章立てされた本格的な【Excelマニュアル】を作成します。")

if not API_KEY:
    st.error("システム設定エラー: Renderの環境変数に GEMINI_API_KEY が設定されていません。")
    st.stop()

uploaded_video = st.file_uploader("マニュアル化する動画をアップロード (MP4/MOV)", type=["mp4", "mov"])

if st.button("Excelマニュアルを作成する", type="primary"):
    if not uploaded_video:
        st.warning("動画ファイルを先にアップロードしてください。")
    else:
        with st.spinner("AIが動画を解析し、Excelを作成中...（数分かかります）"):
            try:
                # 準備
                genai.configure(api_key=API_KEY)
                temp_dir = tempfile.mkdtemp()
                
                video_path = os.path.join(temp_dir, "temp_video.mp4")
                with open(video_path, "wb") as f:
                    f.write(uploaded_video.read())

                st.info("AIに動画を送信しています...")
                video_file = genai.upload_file(path=video_path)
                
                while video_file.state.name == "PROCESSING":
                    time.sleep(5)
                    video_file = genai.get_file(video_file.name)

                # ⭐ ここが超重要：AIへのプロンプトを階層構造に変更
                st.info("AIが内容を分析し、マニュアルを執筆しています...")
                model = genai.GenerativeModel('gemini-2.5-flash')
                prompt = """
                この動画を解析して、操作マニュアルを日本語で詳しく作成してください。
                動画全体の流れを「章（セクション）」に分け、各章の中に具体的な「ステップ」を記述する階層構造にしてください。
                必ず以下のJSON形式のフォーマットのみで出力してください。Markdown(```json等)は一切含めないでください。

                {
                  "title": "マニュアルのタイトル（例：GmailのAIアシスタント機能 活用マニュアル）",
                  "description": "マニュアル全体の概要説明",
                  "sections": [
                    {
                      "heading": "1. はじめに",
                      "time_range": "00:00 - 00:30",
                      "summary": "この章の概要や目的の説明",
                      "steps": [
                        {"time": "00:10", "text": "〇〇の画面を開きます。"}
                      ]
                    },
                    {
                      "heading": "2. 新規メールの作成手順",
                      "time_range": "00:31 - 01:52",
                      "summary": "ゼロから新しいメールを作成する際の手順です。",
                      "steps": [
                        {"time": "00:40", "text": "新規メール作成画面を開きます。"},
                        {"time": "00:50", "text": "「文書作成サポート」ボタンをクリックします。"}
                      ]
                    }
                  ]
                }
                """
                response = model.generate_content([prompt, video_file])
                
                # セキュリティ：即座に動画削除
                genai.delete_file(video_file.name)
                
                # JSON抽出
                ai_data = extract_json_from_text(response.text)

                # ==========================================
                # Excel文書の作成スタート
                # ==========================================
                st.info("画像を抽出してExcelファイルに整理しています...")
                
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "業務マニュアル"
                
                font_meiryo = Font(name='メイリオ')
                font_meiryo_bold = Font(name='メイリオ', bold=True)
                
                # 1行目：全体のタイトル（A〜C列を結合）
                ws.merge_cells('A1:C1')
                ws['A1'] = ai_data.get('title', 'AI自動生成 操作マニュアル')
                ws['A1'].font = Font(name='メイリオ', size=16, bold=True)
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 40
                
                # 2行目：全体の概要（A〜C列を結合）
                ws.merge_cells('A2:C2')
                ws['A2'] = ai_data.get('description', '')
                ws['A2'].font = font_meiryo
                ws['A2'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[2].height = 80
                
                # 3行目：ヘッダー（見出し）
                headers = ['STEP / 時間', '操作説明', '画面画像']
                header_fill = PatternFill(patternType='solid', fgColor='D9D9D9') # 薄いグレー
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col_num, value=header)
                    cell.font = font_meiryo_bold
                    cell.alignment = Alignment(horizontal='center')
                    cell.fill = header_fill

                # 列の幅を設定
                ws.column_dimensions['A'].width = 15
                ws.column_dimensions['B'].width = 60
                ws.column_dimensions['C'].width = 120

                current_row = 4
                step_counter = 1 # マニュアル全体での通し番号
                
                # 章見出し用の背景色（青色）
                section_fill = PatternFill(patternType='solid', fgColor='4F81BD')
                
                # 各章（セクション）ごとの処理
                for section in ai_data.get('sections', []):
                    # ① 章のタイトル（青背景で白文字）
                    ws.merge_cells(f'A{current_row}:C{current_row}')
                    heading_text = f"{section.get('heading', '')} [{section.get('time_range', '')}]"
                    cell = ws.cell(row=current_row, column=1, value=heading_text)
                    cell.font = Font(name='メイリオ', size=14, bold=True, color='FFFFFF')
                    cell.fill = section_fill
                    cell.alignment = Alignment(vertical='center')
                    ws.row_dimensions[current_row].height = 30
                    current_row += 1
                    
                    # ② 章の概要説明
                    summary = section.get('summary', '')
                    if summary:
                        ws.merge_cells(f'A{current_row}:C{current_row}')
                        cell = ws.cell(row=current_row, column=1, value=summary)
                        cell.font = font_meiryo
                        cell.alignment = Alignment(wrap_text=True, vertical='top')
                        ws.row_dimensions[current_row].height = 60
                        current_row += 1
                        
                    # ③ 具体的なステップ（既存の処理）
                    for step in section.get('steps', []):
                        time_str = step.get('time', '')
                        desc_text = step.get('text', '')
                        img_path = os.path.join(temp_dir, f"frame_{current_row}.jpg")
                        
                        cell_a = ws.cell(row=current_row, column=1, value=f"STEP {step_counter}\n({time_str})")
                        cell_a.font = font_meiryo_bold
                        cell_a.alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
                        
                        cell_b = ws.cell(row=current_row, column=2, value=desc_text)
                        cell_b.font = font_meiryo
                        cell_b.alignment = Alignment(vertical='top', wrap_text=True)
                        
                        if extract_frame(video_path, time_str, img_path):
                            img = ExcelImage(img_path)
                            original_width = img.width
                            original_height = img.height
                            
                            target_width = 800
                            if original_width > 0:
                                target_height = int(original_height * (target_width / original_width))
                                img.width = target_width
                                img.height = target_height
                                ws.row_dimensions[current_row].height = (target_height * 0.75) + 15
                            else:
                                ws.row_dimensions[current_row].height = 200
                            ws.add_image(img, f'C{current_row}')
                        else:
                            ws.row_dimensions[current_row].height = 60
                            
                        current_row += 1
                        step_counter += 1
                
                # Excelファイルの保存
                excel_output_path = os.path.join(temp_dir, "AI_Manual_Result.xlsx")
                wb.save(excel_output_path)
                st.success("全行程が完了しました！")

                # ダウンロードボタン
                with open(excel_output_path, "rb") as f:
                    st.download_button(
                        label=" Excelマニュアルをダウンロード",
                        data=f,
                        file_name="Business_Manual.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

            except Exception as e:
                st.error(f"処理中にエラーが発生しました: {e}")
